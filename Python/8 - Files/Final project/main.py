# Projeto para uma loja/supermercado etc...
import os
os.system("pip install openpyxl") # Instalação de Livrarias
import openpyxl
import random
from datetime import date
ddmmyy = date.today().strftime('%d-%m-%Y') # dia mes e ano atual
import filemanagement as f # TRATAMENTO DE FICHEIROS

def listar(l,w): # lista os dados list de forma sequencial
    if l != []:
        print('{}: '.format(w),end='')
        print(*l, sep=', ', end='.')
        print('')
    else:
        print('Sem dados de',w.lower())

def RelatórioVendas(data): # Cria um Relatório de Vendas em Excel
    book = openpyxl.Workbook()
    book.create_sheet('Vendas')
    book.remove(book['Sheet'])
    book['Vendas'].column_dimensions['A'].width = 17
    book['Vendas'].column_dimensions['B'].width = 22
    book['Vendas'].column_dimensions['C'].width = 13
    book['Vendas'].column_dimensions['D'].width = 17
    book['Vendas'].append(['Código EAN-13', 'Nome do Produto', 'Quantidade', 'Preço'])
    total = 0
    for code in vendas_diarias[data]:
        dadosVenda = vendas_diarias[data][code]
        dadosProduto = produtos[code]
        total += dadosVenda[1]
        book['Vendas'].append([str(code), dadosProduto[0], dadosVenda[0], str(dadosVenda[1])+'€'])
    book['Vendas'].append(['Total', str(total)+'€'])
    Name = 'Vendas_'+data+'.xlsx'
    book.save(Name)

def RelatórioEncomendas(): # Cria um Relatório de Encomendas em Excel
    book = openpyxl.Workbook()
    book.create_sheet('Encomendas')
    book.remove(book['Sheet'])
    book['Encomendas'].column_dimensions['A'].width = 20
    book['Encomendas'].column_dimensions['B'].width = 15
    book['Encomendas'].column_dimensions['C'].width = 35
    book['Encomendas'].column_dimensions['D'].width = 20
    book['Encomendas'].column_dimensions['E'].width = 30
    book['Encomendas'].column_dimensions['F'].width = 12
    book['Encomendas'].append(['Fornecedor', 'Tel.', 'Email', 'Código EAN-13', 'Nome do Produto', 'Quantidade'])
    for code in produtos:
        dadosProduto = produtos[code]
        if dadosProduto[1]['Q_Armazém'] + dadosProduto[1]['Q_Exposto'] < dadosProduto[1]['DEFReposição'][0]:
            book['Encomendas'].append([dadosProduto[1]['Fornecedor'], fornecedores[dadosProduto[1]['Fornecedor']][0], fornecedores[dadosProduto[1]['Fornecedor']][1], str(code), dadosProduto[0],  dadosProduto[1]['DEFReposição'][0]])
    book.save('Encomendas.xlsx')

def menu():
    print('0 - Sair')
    print('1 - Consultar Informação') 
    print('2 - Adicionar/Alterar Informação') 
    print('3 - Remover Informação')
    print('4 - Ler produtos (Realizar uma comprar)')
    print('5 - Gerar Relatórios')
    print('6 - Salvar dados')

produtos = {} # int(CódigoProduto) : list("NomeProduto", dict("Preço":int(preço), "Categoria":'categoria', "Q_Armazém": int(q_a), "Q_Exposto"): int(q_e), "DEFReposição":list(Q_MINrepor, Q_Encomenda), "Fornecedor":'NomeFornecedor')
categorias = {} # "Categoria" : list(int(CódigoProduto1), int(CódigoProduto2))
fornecedores = {} # "NomeFornecedor" : list(int(Tel), 'Email')
vendas_diarias = {ddmmyy:{}} # data : {int(CódigoProduto1): [int(quantidade), int(preço*quantidade]}
caixaRegis = [] # list(int(CódigoProduto1), Quant, Preço*Quant), list(int(CódigoProduto1), Quant, Preço*Quant) estrutura TEMP para caso a pessoa deixe a compra pendente

f.downloadProdutos(produtos)
f.downloadCategorias(categorias)
f.downloadFornecedores(fornecedores) # BAIXA PARA A MEMÓRIA TODOS OS DADOS SALVOS
f.downloadVendas(vendas_diarias)

for data in vendas_diarias: # Cria o relatório de vendas quando o programa for aberto no dia seguinte
    if data != ddmmyy:
        RelatórioVendas(data)
        vendas_diarias.pop(data)

sep = '----------'
print(sep)
print('Programa de Gestão')
print(sep)
while True:
    menu()
    acao = int(input('- '))
    print(sep)
    if acao == 0:
        while acao < 1 or acao > 3:
            print('Pretende salvar os dados?\n1 - Sim\n2 - Não\n3 - Cancelar')
            acao = int(input('- '))
        if acao == 1:
            f.uploudProdutos(produtos)
            f.uploudCategorias(categorias)
            f.uploudFornecedores(fornecedores) # SALVA TODOS OS DADOS EM MEMÓRIA
            f.uploudVendas(vendas_diarias)
            print(sep)
            exit()
        elif acao == 2:
            print(sep)
            exit()
        else:
            continue
    elif acao == 1:
        acao = -1 # NECESSÁRIO PARA ENTRAR NA CONDIÇÃO WHILE
        while acao < 0 or acao > 4:
            print('0 - Voltar')
            print('1 - Consultar Produtos')
            print('2 - Consultar Categorias')
            print('3 - Consultar Fornecedores')
            print('4 - Nº de vendas e saldo total diário')
            acao = int(input('- '))
            print(sep)
        if acao == 0:
            continue
        if acao == 1:
            acao = -1 
            while acao < 0 or acao > 2:
                print('0 - Voltar')
                print('1 - Por categorias')
                print('2 - Por código EAN-13') # Código EAN-13 == Código de Barras
                acao = int(input('- '))
                print(sep)
            if acao == 0:
                continue
            elif acao == 1:
                listar(list(categorias), 'Categorias')
                cat = input('Escolha uma categoria: ')
                if cat in categorias and categorias[cat] != []:
                    produtosCAT = categorias[cat] # Todos os códigos de produtos na categoria
                    for code in produtosCAT:
                        dadosProduto = produtos[code] 
                        print('Código EAN-13: {0} | {1} | Preço: {2}€ | Armazém: {3} uni. | Exposto: {4} uni. \nRepor q. houver: {5} uni. | Encomenda Pred.: {6} uni. | Fornecedor: {7}'.format(code, dadosProduto[0], dadosProduto[1]['Preço'], dadosProduto[1]['Q_Armazém'], dadosProduto[1]['Q_Exposto'], dadosProduto[1]['DEFReposição'][0], dadosProduto[1]['DEFReposição'][1], dadosProduto[1]['Fornecedor']))
                else:
                    print('A categoria inserida não consta na base de dados, ou não tem produtos associados')
            else:
                code = int(input('Introduza o código EAN-13 do produto: '))
                while len(str(code)) != 13:
                    print('O código EAN-13 tem de ter 13 digitos')
                    code = int(input('Introduza o código EAN-13 do produto: '))
                if code in produtos:
                    dadosProduto = produtos[code]
                    print('{0} | Preço: {1}€ | Categoria: {2} | Armazém: {3} uni. | Exposto: {4} uni. \nRepor q. houver: {5} uni. | Encomenda Pred.: {6} uni. | Fornecedor: {7}'.format(dadosProduto[0], dadosProduto[1]['Preço'], dadosProduto[1]['Categoria'], dadosProduto[1]['Q_Armazém'], dadosProduto[1]['Q_Exposto'], dadosProduto[1]['DEFReposição'][0], dadosProduto[1]['DEFReposição'][1], dadosProduto[1]['Fornecedor']))
                else:
                    print('O código EAN-13 inserido não consta na base de dados.')
            print(sep)
        elif acao == 2:
            listar(list(categorias), 'Categorias')
            print(sep)
        elif acao == 3:
            listar(list(fornecedores), 'Fornecedores')
            fornecedor = input('Introduza o nome do fornecedor: ')
            while fornecedor not in fornecedores:
                print(fornecedor,'não se encontra na base de dados.')
                fornecedor = input('Introduza o nome do fornecedor novamente: ')
            print('Tel: {0} | Email: {1}'.format(fornecedores[fornecedor][0],fornecedores[fornecedor][1]))
        else:
            vendas = 0
            rendimento = 0
            for code in vendas_diarias[ddmmyy]: # Verifica o rendimento e as vendas de hoje
                vendas += vendas_diarias[ddmmyy][code][0]
                rendimento += produtos[code][1]['Preço']*vendas_diarias[ddmmyy][code][1]
            print('Totais de vendas (Hoje):',vendas)
            print('Rendimento Total (Hoje): {}€'.format(round(rendimento,2)))
    elif acao == 2:
        acao = -1
        while acao < 0 or acao > 3:
            print('0 - Voltar')
            print('1 - Adicionar/Alterar Produtos')
            print('2 - Adicionar/Alterar Categorias')
            print('3 - Adicionar/Alterar Fornecedores')
            acao = int(input('- '))
            print(sep)
        if acao == 0:
            continue
        elif acao == 1:
            code = -1
            while len(str(code)) != 13 and code != 0:
                print('0 - Gerar código aleatório')
                code = int(input('Introduza o código do produto (EAN-13 dígitos): '))
            if code == 0: # Em caso de 0 ele próprio gera o código EAN-13
                code = random.randint(1000000000000, 9999999999999)
                while code in produtos:
                    code = random.randint(1000000000000, 9999999999999)
            if code in produtos:  ########################ATUALIZAÇÃO########################
                print('O código EAN 13 - {} já existe.'.format(code))
                acao = -1
                while acao < 0 or acao > 7:
                    print('O que pretende alterar:')
                    print(' 0 - Voltar')
                    print(' 1 - Nome do Produto')
                    print(' 2 - Preço')
                    print(' 3 - Categoria')
                    print(' 4 - Quantidade em Armazém')
                    print(' 5 - Quantidade em Exposto')
                    print(' 6 - Definições de Reposição/Encomenda')
                    print(' 7 - Fornecedor')
                    acao = int(input('- '))
                if acao == 0:
                    continue
                elif acao == 1:
                    nome = input('Introduza o novo nome do produto: ')
                    Nexiste = True
                    while Nexiste: # Proteções da alteração do nome
                        for codigo in produtos:
                            if nome == produtos[codigo][0]:
                                Nexiste = False
                        if Nexiste == False:
                            Nexiste = True
                            print('O nome já é conhecido na base de dados tente usar outro.')
                            nome = input('Introduza o novo nome do produto: ')
                        else:
                            Nexiste = False
                    produtos[code][0] = nome # Atualização do nome
                    print('Produto atualizado com sucesso.')
                elif acao == 2:
                    preço = round(float(input('Introduza o novo preço do produto ex."9.99": ')),2)
                    while preço < 0: # Proteção da alteração do preço
                        print('O preço do produto não pode ser negativo.')
                    produtos[code][1]['Preço'] = preço # Atualização do preço
                    print('Produto atualizado com sucesso.')
                elif acao == 3:
                    listar(list(categorias), 'Categorias')
                    categoria = input('Introduza o nome da categoria em que o produto se insere: ')
                    while categoria not in categorias: # Proteção de categoria
                        print(categoria,'não se encontra na base de dados.')
                        categoria = input('Introduza o nome da categoria em que o produto se insere: ')
                    categoriaANT = produtos[code][1]['Categoria'] # Categoria antiga do produto
                    categorias[categoriaANT].pop(categorias[categoriaANT].index(code)) # No dict categorias na categoria Antiga remove o produto 
                    produtos[code][1]['Categoria'] = categoria # Atualização a categoria
                    categorias[categoria].append(code) # Adiciona no dict categoria na categoria nova o produto
                    print('Produto atualizado com sucesso.')
                elif acao == 4:
                    q_armazém = int(input('Introduza as unidades existentes em armazém: '))
                    while q_armazém < 0:
                        print('O valor das unidades existentes em armazém tem de ser positivas ou nulas.')
                        q_armazém = int(input('Introduza as unidades existentes em armazém: '))
                    produtos[code][1]['Q_Armazém'] = q_armazém # Atualização do stock em armazém
                    print('Produto atualizado com sucesso.')
                elif acao == 5:
                    q_exposto = int(input('Introduza as unidades já expostas em loja: '))
                    while q_exposto < 0:
                        print('O valor das unidades já expostas em loja tem de ser positivas ou nulas.')
                        q_exposto = int(input('Introduza as unidades existentes em loja: '))
                    produtos[code][1]['Q_Exposto'] = q_armazém # Atualização do stock exposto
                    print('Produto atualizado com sucesso.')
                elif acao == 6:
                    q_MINrepor = int(input('Introduza as unidades mínimas em loja&armazém para acionar a encomenda: '))
                    while q_MINrepor < 1:
                        print('O valor das unidades mínimas em loja&armazém tem de ser positivas.')
                        q_MINrepor = int(input('Introduza as unidades mínimas em loja&armazém para acionar a encomenda: '))
                    produtos[code][1]['DEFReposição'][0] = q_MINrepor # # Atualização das unidades mínimas em loja&armazém
                    q_Encomenda = int(input('Introduza as unidades a encomendar predefinidas: '))
                    while q_Encomenda < 1:
                        print('O valor das unidades a encomendar tem de ser positivas.')
                        q_Encomenda = int(input('Introduza as unidades a encomendar predefinidas: '))
                    produtos[code][1]['DEFReposição'][1] = q_Encomenda # Atualização unidades a encomendar predefinidas
                    print('Produto atualizado com sucesso.')
                else:
                    listar(list(fornecedores), 'Fornecedores')
                    fornecedor = input('Introduza o nome do fornecedor: ')
                    while fornecedor not in fornecedores:
                        print(fornecedor,'não se encontra na base de dados.')
                        fornecedor = input('Introduza o nome do fornecedor: ')
                    produtos[code][1]['Fornecedor'] = fornecedor # Atualização do Fornecedor
                    print('Produto atualizado com sucesso.')
            else:   ########################ADIÇÃO########################
                print('---NOVO PRODUTO---')
                nome = input('Introduza o nome do novo produto: ')
                Nexiste = True
                while Nexiste: # Proteções da adição do nome
                    for codigo in produtos:
                        if nome == produtos[codigo][0]:
                            Nexiste = False
                    if Nexiste == False:
                        Nexiste = True
                        print('O nome já é conhecido na base de dados tente usar outro.')
                        nome = input('Introduza o nome do novo produto: ')
                    else:
                        Nexiste = False 
                preço = round(float(input('Introduza o preço do novo produto ex."9.99": ')),2)
                while preço < 0: # Proteção da adição do preço
                    print('O preço do produto não pode ser negativo.')
                listar(list(categorias), 'Categorias')
                categoria = input('Introduza o nome da categoria em que o produto se insere: ')
                while categoria not in categorias: # Proteção de categoria
                    print(categoria,'não se encontra na base de dados.')
                    categoria = input('Introduza o nome da categoria em que o produto se insere: ')
                q_armazém = int(input('Introduza as unidades existentes em armazém: '))
                while q_armazém < 0:
                    print('O valor das unidades existentes em armazém tem de ser positivas ou nulas.')
                    q_armazém = int(input('Introduza as unidades existentes em armazém: '))
                q_exposto = int(input('Introduza as unidades já expostas em loja: '))
                while q_exposto < 0:
                    print('O valor das unidades já expostas em loja tem de ser positivas ou nulas.')
                    q_exposto = int(input('Introduza as unidades já expostas em loja: '))
                q_MINrepor = int(input('Introduza as unidades mínimas em loja&armazém para acionar a encomenda: '))
                while q_MINrepor < 1:
                    print('O valor das unidades mínimas em loja&armazém tem de ser positivas.')
                    q_MINrepor = int(input('Introduza as unidades mínimas em loja&armazém para acionar a encomenda: '))
                q_Encomenda = int(input('Introduza as unidades a encomendar predefinidas: '))
                while q_Encomenda < 1:
                    print('O valor das unidades a encomendar tem de ser positivas.')
                    q_Encomenda = int(input('Introduza as unidades a encomendar predefinidas: '))
                fornecedor = input('Introduza o nome do fornecedor: ')
                while fornecedor not in fornecedores:
                    print(fornecedor,'não se encontra na base de dados.')
                    fornecedor = input('Introduza o nome do fornecedor novamente: ')
                produtos[code] = [nome]
                produtos[code].append({"Preço":preço})
                produtos[code][1]['Categoria'] = categoria
                categorias[categoria].append(code)
                produtos[code][1]['Q_Armazém'] = q_armazém #######ADIÇÕES#######
                produtos[code][1]['Q_Exposto'] = q_exposto
                produtos[code][1]['DEFReposição'] = [q_MINrepor]
                produtos[code][1]['DEFReposição'].append(q_Encomenda)
                produtos[code][1]['Fornecedor'] = fornecedor
                print('Produto adicionado com sucesso.')
        elif acao == 2:
            listar(list(categorias),'Categorias')
            categoria = input('Introduza o nome da categoria: ')
            if categoria in categorias:
                categoriaANT = categoria # Categoria Antiga
                print('A categoria - {} já existe.'.format(categoria))
                info = categorias[categoria] # Todos os produtos da categoria Antiga
                categoria = input('Introduza o novo nome a alterar: ') # Novo nome da Categoria
                while categoria in categorias: # Proteção de categoria pois não pode ser repetido
                    print(categoria,'encontra se na base de dados.')
                    categoria = input('Introduza o novo nome a alterar: ')
                categorias.pop(categoriaANT) # Remove a antiga
                categorias[categoria] = info # Adiciona a Nova com a info da antiga (Atualização)
                for code in info: 
                    produtos[code][1]['Categorias'] = categoria # Substitui todos os produtos com a Categoria antiga para a nova
                print('Categoria renomeada com sucesso.')
            else:
                categorias[categoria] = [] # Cria nova se não existir
                print('Nova categoria criada com sucesso.')
        else:
            listar(list(fornecedores),'Fornecedores')
            fornecedor = input('Introduza o nome do fornecedor: ')
            if fornecedor in fornecedores:
                print('O fornecedor - {} já existe.'.format(categoria))
                Nfornecedor = input('Introduza o novo nome do fornecedor: ')
                while Nfornecedor in fornecedores: # Proteção Nfornecedor
                    print('O novo nome não pode ser conhecido.')
                    Nfornecedor = input('Introduza o novo nome do fornecedor a alterar: ')
                tel = int(input('Introduza o novo telefone do fornecedor: '))
                while len(str(tel)) < 9 or len(str(tel)) > 12: # proteção TEL
                    print('O nº de Tel não pode ter',len(str(tel),'digitos.'))
                    tel = int(input('Introduza o novo telefone do fornecedor: '))
                email = input('Introduza o novo email do fornecedor: ')
                while '@' not in email: # Proteção EMAIL
                    print('O email tem de estar com o devido @.')
                fornecedores.pop(fornecedor) # REMOVE O ANTIG 
                fornecedores[Nfornecedor] = [tel,email] # Adiciona o novo
                for code in produtos:
                    if produtos[code][1]['Fornecedor'] == fornecedor:  # Altera os produtos com o com fornecedor antigo
                        produtos[code][1]['Fornecedor'] = Nfornecedor
                print('Fornecedor atualizada com sucesso.')
            else:
                tel = int(input('Introduza o novo telefone do fornecedor: '))
                email = input('Introduza o novo email do fornecedor: ')
                fornecedores[fornecedor] = [tel,email] # CASO não exista adiciona
                print('Fornecedor criado com sucesso.')
    elif acao == 3:
        acao = -1
        while acao < 0 or acao > 3:
            print('0 - Voltar')
            print('1 - Remover Produtos')
            print('2 - Remover Categorias')
            print('3 - Remover Fornecedores')
            acao = int(input('- '))
        if acao == 0:
            continue
        elif acao == 1:
            code = int(input('Introduza o código do produto a remover: '))
            while len(str(code)) != 13: # Proteção Código EAN-13
                print('O código EAN-13 tem de ter 13 digitos')
                code = int(input('Introduza o código do produto a remover: '))
            if code in produtos:
                caixaRegis.clear() # remover o produto da caixa registadora
                if code in vendas_diarias[ddmmyy]: # ddmmyy = hoje
                    vendas_diarias[ddmmyy].pop(code) # remover o produto das vendas diárias 
                for cat in categorias: # remover o produto das categorias
                    if code in categorias[cat]:
                        categorias[cat].pop(categorias[cat].index(code))
                produtos.pop(code) # remover o produto
                print('Produto removido com sucesso.')
            else:
                print('Código introduzido não consta na base de dados.')
        elif acao == 2:
            listar(list(categorias),'Categorias')
            categoria = input('Introduza a categoria a remover: ')
            if categoria in categorias:
                for code in categorias[categoria]: # itera todos os produtos que vão ficar sem categoria
                    listar(list(categorias),'Categorias')
                    print('O produto {} está sem categoria.'.format(produtos[code][0]))
                    cat = input('Introduza o nome da categoria em que o produto se insere: ')
                    while cat not in categorias: # Proteção de categoria
                        print(cat,'não se encontra na base de dados.')
                        cat = input('Introduza o nome da categoria em que o produto se insere: ')
                    categorias[cat].append(code) # Adiciona todos os produtos que vão ficar sem categoria às suas novas categorias
                    produtos[code][1]['Categoria'] = cat # Altera em produtos a sua categoria
                categorias.pop(categoria) # Remove a categoria
                print('Categoria removida com sucesso.')
            else:
                print('Categoria introduzida não consta na base de dados.')
        elif acao == 3:
            listar(list(fornecedores),'Fornecedores')
            fornecedor = input('Introduza o fornecedor a remover: ')
            if fornecedor in fornecedores:
                for code in produtos:
                    if produtos[code][1]['Fornecedor'] == fornecedor:
                        listar(list(fornecedores),'Fornecedores')
                        print('O produto {} está sem Fornecedor.'.format(produtos[code][0]))
                        forn = input('Introduza o nome do Fornecedor do produto: ')
                        while forn not in fornecedores: # Proteção de categoria
                            print(forn,'não se encontra na base de dados.')
                            forn = input('Introduza o nome do Fornecedor do produto: ')
                        produtos[code][1]['Fornecedor'] = forn # Adiciona todos os produtos que vão ficar sem fornecedor aos seus novos fornecedores
                fornecedores.pop(fornecedor) # Remove o fornecedor
                print('Fornecedor removido com sucesso.')
            else:
                print('Forncedor introduzido não consta na base de dados.')
    elif acao == 4:
        while True:
            acao = -1
            while acao < 0 or acao > 2:
                listar(caixaRegis,'Itens') # A melhorar
                print('0 - Voltar')
                print('1 - Passar Produto')
                print('2 - Confirmar Pagamento')
                print('3 - Limpar Compra')
                acao = int(input('- '))
                print(sep)
            if acao == 0:
                
                break
            elif acao == 1:
                code = int(input('Introduza o código EAN-13 do item: '))
                while len(str(code)) != 13:
                    print('O código EAN-13 tem de ter 13 digitos')
                    code = int(input('Introduza o código EAN-13 do item: '))
                if code in produtos:
                    v = False
                    for i in range(len(caixaRegis)):
                        if code in caixaRegis[i]:
                            caixaRegis[i][1] += 1 # Dá mais 1 à quantidade
                            caixaRegis[i][2] += produtos[code][1]['Preço'] # Dá mais um preço
                            v = True
                    if v == False:
                        caixaRegis.append([code, 1, produtos[code][1]['Preço']]) # se n existir adiciona
                else:
                    print('Código introduzido não consta na base de dados.')
            elif acao == 2:
                for i in caixaRegis:
                    if produtos[i[0]][1]['Q_Exposto'] > 0:
                        produtos[i[0]][1]['Q_Exposto'] -= i[1] # Retira aos produtos expostos
                        d = vendas_diarias[ddmmyy].get(i[0],[0,0])
                        vendas_diarias[ddmmyy][i[0]] = [d[0]+i[1],d[1]+i[2]] # atualiza as vendas diarias
                    else:
                        print(produtos[i[0]][0],'está esgotado.')
                caixaRegis.clear()
                break
            else:
                caixaRegis.clear()
    elif acao == 5:
        # Relatório Encomendas para Fornecederes
        RelatórioEncomendas()
        # Relatório Vendas Diárias
        RelatórioVendas(ddmmyy)
        print('Relatórios em', os.getcwd())
    elif acao == 6:
        f.uploudProdutos(produtos)
        f.uploudCategorias(categorias)
        f.uploudFornecedores(fornecedores) # SALVA TODOS OS DADOS EM MEMÓRIA
        f.uploudVendas(vendas_diarias)